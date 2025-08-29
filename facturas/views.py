from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.detail import DetailView
from django.views.generic.list import ListView
from django.views.generic.base import TemplateView
from .models import Facturas, FacturasTransactions, Productos, Clientes, ClienteContactos, ClienteDireccion
from .forms import Facturas_Form, ProductoForm, ClienteDireccionForm, ClienteForm, ContactoForm, FacturasTransactionsForm
from django.urls import reverse_lazy, reverse
from django.http import HttpResponse
from django.template.loader import get_template

from django.shortcuts import redirect, get_object_or_404, render
from django.utils import timezone
from django.forms import modelformset_factory
from django.contrib import messages
from django.db import transaction  # Importación faltante


#estas dos importaciones es para saber si el usuario es miembro del staff y evitarnos la condicion del mixin
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required #IMPORTANTE: me permite saber si el usuario se registro

# Create your views here.


class StaffRequiredMixin(object): #Este mixin requerira que el usuario sea miembro del staff
    @method_decorator(staff_member_required) #este decorador es para verificar si el usuario es staff o no
    def dispatch(self, request, *args, **kwargs):
        print(request.user)
        # if not request.user.is_staff: #esto es para saber si el usuario ha iniciado sesion en el admin, si no ha iniciado te redirecciona al admin
        #     return redirect(reverse_lazy("admin:login"))
        return super(StaffRequiredMixin, self).dispatch(request, *args, **kwargs) #el dispatch procesa la peticiones al servidor
    

@method_decorator(login_required, name='dispatch') #este decorador es para verificar si el usuario es staff y asi poder usar estas funcionalidades
class FacturaCreateView(CreateView):
    model = Facturas
    form_class = Facturas_Form
    # template_name = 'facturas/factura_form.html'  # Asegúrate de que esta plantilla exista
    
            
    def get_success_url(self):
        return reverse_lazy('facturas:facturas_list')
  
 

    def form_invalid(self, form):
        """
        Renderiza la vista con los errores del formulario.
        """
        context = self.get_context_data()
        messages.error(self.request, "Hubo un error al crear la factura. Por favor, revisa los datos.")
        return self.render_to_response(context)
    
  


    
@method_decorator(login_required, name='dispatch')
class FacturaListView(ListView):
    model=Facturas
    context_object_name = 'factura_list'
    template_name = 'factura_list.html'
    paginate_by=5
    
    def get_queryset(self): #ordena la factura por numero de factura
        return Facturas.objects.all().order_by('-invoice_n')
  
           
    
    
@method_decorator(login_required, name='dispatch')
class FacturaUpdateView(UpdateView):
    model = Facturas
    form_class=Facturas_Form
    # fields = ["descripcion", "total"]
    template_name_suffix = "_update_form"
    
    def get_success_url(self):
      
        return reverse_lazy('facturas:facturas_list')
    
@method_decorator(login_required, name='dispatch')

class FacturaDeleteView(DeleteView):
    model = Facturas
    success_url = reverse_lazy('facturas:facturas_list')
    

class FacturaDetailView(DetailView):
    model = Facturas
    context_object_name = 'Facturas'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        factura = self.object  # Obtiene la factura actual
        total_balance = sum(transaction.calcular_total() for transaction in FacturasTransactions.objects.filter(invoice_id=factura))
        context['total_balance'] = total_balance
        return context
    

#Vistas de transacciones

@method_decorator(login_required, name='dispatch')
class  FacturasTransactionsCreateView(CreateView):
    model = FacturasTransactions
    form_class = FacturasTransactionsForm
    template_name = 'facturas/facturas_transaction_form.html'
    
    
    def get_success_url(self):
        return reverse_lazy('facturas:facturas_list')
    
    def form_valid(self, form):
        factura = get_object_or_404(Facturas, pk=self.kwargs['product_id'])  # Obtén el objeto relacionado
        form.instance.invoice_id = factura  # Asocia el `partner_id` al formulario
        return super().form_valid(form)
    
    
    

@method_decorator(login_required, name="dispatch")
class FacturasTransactionsUpdateView(UpdateView):
     model = FacturasTransactions
     form_class = FacturasTransactionsForm
     template_name_suffix = "_update_form"
     template_name = 'facturas/facturas_transaction_update_form.html'
     
    #  def get_success_url(self):
    #     return reverse_lazy('facturas:facturas_list')
    
     def get_success_url(self):
        factura_id = self.object.invoice_id.id  # Obtener el ID de la factura asociada
        return reverse('facturas:facturas_detail', kwargs={'pk': factura_id})
    
@method_decorator(login_required, name="dispatch")

class FacturasTransactionDeleteView(DeleteView):
    model = FacturasTransactions
    template_name = 'facturas/facturas_transaction_confirm_delete.html'
    # success_url = reverse_lazy('facturas:facturas_list')
    def get_success_url(self):
        factura_id = self.object.invoice_id.id  # Obtener el ID de la factura asociada
        return reverse('facturas:facturas_detail', kwargs={'pk': factura_id})
    
    
    
    
#Vistas de Productos

#CRUD productos

@method_decorator(login_required, name='dispatch')
class ProductoCreateView(CreateView):
    model = Productos
    form_class = ProductoForm
    
    def get_success_url(self):
        return reverse_lazy('productos:productos_list')
    

@method_decorator(login_required, name='dispatch')
class ProductoListView(ListView):
    model = Productos
    paginate_by = 100  # if pagination is desired
    context_object_name = 'producto_list'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["now"] = timezone.now()
        return context
    
    
@method_decorator(login_required, name='dispatch')
class ProductoUpdateView(UpdateView):
    model = Productos
    form_class = ProductoForm
    template_name_suffix = "_update_form"
    
    def get_success_url(self):
        return reverse_lazy('productos:productos_list')

@method_decorator(login_required, name='dispatch')

class ProductoDeleteView(DeleteView):
    model = Productos
    success_url = reverse_lazy('productos:productos_list')
    

@method_decorator(login_required, name='dispatch')
class ProductoDetailView(DetailView):
    model = Productos
    context_object_name="Productos"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["now"] = timezone.now()
        return context


#Vistas de Clientes
@method_decorator(login_required, name='dispatch') 
class ClienteCreateView(CreateView):
    model = Clientes
    # fields = ["nombre", "rif", "telefono_1", "telefono_2", "email", "num_empleados"]
    form_class=ClienteForm
    template_name = 'facturas/Clientes_form.html'
    
    def get_success_url(self):
      
        return reverse_lazy('cliente:cliente_list')


@method_decorator(login_required, name='dispatch') 
class ClienteListView(ListView):
    model=Clientes
    context_object_name = 'cliente_list'
    template_name = 'facturas/cliente_list.html'
    
@method_decorator(login_required, name='dispatch') 
class ClienteUpdateView(UpdateView):
    model = Clientes
    # fields = ["rif", "nombre", "email", "telefono_1", "telefono_2", "num_empleados"]
    form_class=ClienteForm
    template_name_suffix = "_update_form"
    
    
    def get_success_url(self):
      
        return reverse_lazy('cliente:cliente_list')
    
    
@method_decorator(login_required, name='dispatch') 
class ClienteDeleteView(DeleteView):
    model = Clientes
    success_url = reverse_lazy('cliente:cliente_list')
    
@method_decorator(login_required, name='dispatch') 
class ClienteDetailView(DetailView):
    model = Clientes
    context_object_name = 'partner' #con este atributo puedo usar este nombre para acceder al modelo de cada cliente en particular
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["now"] = timezone.now()
        return context
    
#CRUD DIRECCION CLIENTE

# Crear dirección
class PartnerAddressCreateView(CreateView):
    model = ClienteDireccion
    form_class=ClienteDireccionForm
    # fields = ['address_type', 'address_lines', 'ref_address', 'country_id', 'state_id', 'city', 'municipality', 'parish', 'postal_code']
    template_name = 'facturas/cliente_direccion_form.html'
    def form_valid(self, form):
        # Obtener el cliente usando el parámetro de la URL
        cliente = get_object_or_404(Clientes, pk=self.kwargs['partner_id'])
        form.instance.partner_id = cliente  # Asigna el cliente a la dirección
        return super().form_valid(form)

    def get_success_url(self):
        # Redirige a la lista de clientes, o la página deseada
        return reverse_lazy('cliente:cliente_list')
    
   
    


# Listar direcciones
class PartnerAddressListView(ListView):
    model = ClienteDireccion
    template_name = 'facturas/cliente_direccion_list.html'
    context_object_name = 'addresses'

# Detalle de dirección
class PartnerAddressDetailView(DetailView):
    model = ClienteDireccion
    template_name = 'facturas/cliente_direccion_detail.html'
    context_object_name = 'partner'
    
    def get_object(self):
        # Maneja el error 404 si el objeto no se encuentra
        return get_object_or_404(ClienteDireccion, pk=self.kwargs['pk'])
    

class PartnerAddressUpdateView(UpdateView):
    model = ClienteDireccion
    form_class=ClienteDireccionForm
    template_name = 'facturas/cliente_direccion_update_form.html'
    
    def get_object(self):
        return get_object_or_404(ClienteDireccion, pk=self.kwargs['pk'])
    
    def get_success_url(self):
      
        return reverse_lazy('cliente:cliente_list')


# Eliminar dirección
class PartnerAddressDeleteView(DeleteView):
    model = ClienteDireccion
    template_name = 'facturas/cliente_direccion_confirm_delete.html'
    def get_object(self):
        # Usa el pk de la URL para buscar el objeto de dirección
        return get_object_or_404(ClienteDireccion, pk=self.kwargs['pk'])

    def get_success_url(self):
        # Redirige a la vista de detalle del cliente
        return reverse('cliente:cliente_detail', kwargs={'pk': self.object.partner_id.id})
    
# CRUD CONTACTO
@method_decorator(login_required, name='dispatch') 
class ContactoCreateView(CreateView):
    model = ClienteContactos
    form_class=ContactoForm
    # template_name = 'facturas/cliente_contactos_form.html'
    
    def get_success_url(self):
      
        return reverse_lazy('cliente:cliente_list')
    
    def form_valid(self, form):
        partner = get_object_or_404(Clientes, pk=self.kwargs['cliente_id'])  # Obtén el objeto relacionado
        form.instance.partner_id = partner  # Asocia el `partner_id` al formulario
        return super().form_valid(form)
    


# @method_decorator(login_required, name='dispatch') 
# class ContactoListView(ListView):
#     model=Contacto
#     context_object_name = 'contacto_list'
#     template_name = 'contacto_list.html'
    
@method_decorator(login_required, name='dispatch') 
class ContactoUpdateView(UpdateView):
    model = ClienteContactos
    # fields = ["name", "cargo", "telephone", "cellphone", "extension", "email"]
    form_class=ContactoForm
    template_name_suffix = "_update_form"
    
    def get_success_url(self):
      
        return reverse_lazy('cliente:cliente_list')
    
    
@method_decorator(login_required, name='dispatch') 
class ContactoDeleteView(DeleteView):
    model =  ClienteContactos
    success_url = reverse_lazy('cliente:cliente_list')
    
    
#crearemos una funcion para mostrar las facturas del cliente
# @method_decorator(login_required, name='dispatch') 

@login_required
def facturas_por_cliente(request, cliente_id):
   cliente = get_object_or_404(Clientes, id=cliente_id)
   facturas = Facturas.objects.filter(partner_id=cliente)
   return render(request, 'facturas/facturas_por_cliente.html', {'cliente': cliente, 'facturas': facturas})

#Filtro el contacto por cliente
@login_required
def contacto_por_cliente(request, cliente_id):
   cliente = get_object_or_404(Clientes, id=cliente_id)
   contactos = ClienteContactos.objects.filter(partner_id=cliente)
   return render(request, 'facturas/contacto_por_cliente.html', {'cliente': cliente, 'contactos': contactos})

@login_required
def direccion_por_cliente(request, cliente_id):
    # Obtener el cliente específico
    cliente = get_object_or_404(Clientes, id=cliente_id)
    
    # Obtener las direcciones asociadas al cliente
    direcciones = ClienteDireccion.objects.filter(partner_id=cliente)
    
    # Renderizar la plantilla con los datos
    return render(request, 'facturas/direccion_por_cliente.html', {'cliente': cliente, 'direcciones': direcciones})






#generar PDF

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from django.http import HttpResponse
from reportlab.lib.units import mm
from decimal import Decimal
# def generate_pdf(request, pk):
#     # Obtener datos de la factura
#     factura = Facturas.objects.get(pk=pk)

#     # Crear respuesta HTTP para el PDF
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="factura_{pk}.pdf"'

#     # Crear el lienzo del PDF
#     pdf = canvas.Canvas(response, pagesize=letter)
#     pdf.setTitle(f"Factura de {factura.partner_id}")

#     # Estilos de texto
#     pdf.setFont("Helvetica-Bold", 16)
#     pdf.drawString(200, 750, f"Factura de {factura.partner_id}")

#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(50, 720, f"Cliente: {factura.partner_id}")
#     pdf.drawString(50, 700, f"Número de factura: {factura.invoice_n}")
#     pdf.drawString(50, 680, f"Número de control: {factura.invoice_c}")
#     pdf.drawString(50, 660, f"Descuento: {factura.discount}%")
#     pdf.drawString(50, 640, f"Tipo de moneda: {factura.currency_id}")
#     pdf.drawString(50, 620, f"Nota pública: {factura.pub_note}")
#     pdf.drawString(50, 600, f"Fecha de la factura: {factura.invoice_d}")

#     # Datos de los productos en tabla
#     data = [["Producto", "Precio", "Cantidad", "IVA", "Total"]]
    
#     for product in factura.get_factura_transaction():
#         data.append([
#             product.product_id,
#             f"{product.price} bs",
#             product.qty,
#             f"{product.vat_rate}%",
#             f"{product.calcular_total()} bs"
#         ])

#     table = Table(data, colWidths=[150, 80, 80, 80, 80])
#     table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#         ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#         ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
#         ('GRID', (0, 0), (-1, -1), 1, colors.black),
#     ]))

#     # Posicionar la tabla
#     table.wrapOn(pdf, 50, 500)
#     table.drawOn(pdf, 50, 550)

#     # Finalizar el PDF
#     pdf.showPage()
#     pdf.save()

#     return response

# MODELO FACTURA 1

# def generate_pdf(request, pk):
#     factura = Facturas.objects.get(pk=pk)
    
#     def balance_subtotal(factura):
#         subtotal = 0
#         for transaction in factura.get_factura_transaction():
#             subtotal += transaction.calcular_subtotal()
#         return subtotal
    
#     def balance_total(factura):
#         total = 0
#         for transaction in factura.get_factura_transaction():
#             total += transaction.calcular_total()
#         return total
    

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="factura_{pk}.pdf"'
#     pdf = canvas.Canvas(response, pagesize=letter)
#     width, height = letter

#     # Encabezado
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(20, 670, "Cliente:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(20, 658, str(factura.partner_id))

#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(20, 640, "RIF:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(20, 628, str(factura.partner_id.tin))

#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(20, 610, "Domicilio Fiscal:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(20, 598, str(factura.partner_id.name))

#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(20, 580, "Contactos:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(20, 568, str(factura.partner_id.website))

#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(480, 670, f"Factura N°: {factura.invoice_n}")
#     pdf.drawString(480, 655, f"Control N°: {factura.invoice_c}")
#     pdf.drawString(480, 640, f"Fecha: {factura.invoice_d.strftime('%d/%m/%Y')}")

#     # Tabla de productos
#     data = [["Producto", "PRECIO", "CANTIDAD", "TOTAL"]]
#     for product in factura.get_factura_transaction():
#        data.append([
#             product.product_id,
#             f"{product.price} $",
#             product.qty,
#             f"{product.calcular_subtotal():.2f} $",
           
#         ])

#     # 📐 Ancho total para centrar
#     colWidths = [200, 130, 130, 115]  # Total: 595
#     # x_position = (page_width - total_width) / 2  # 42.5 → puedes redondear a 42
#     x_position = 20
#     y_position = 480

#         # 🧾 Construcción y estilo de tabla
#     table = Table(data, colWidths=colWidths)
#     table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (-1, 0), colors.white),
#             ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
#             ('FONTSIZE', (0, 0), (-1, 0), 9),  # Tamaño de fuente más pequeño en el encabezado
#             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),      # Negrita
#             ('ALIGN', (0, 0), (-1, 0), 'CENTER'),                 # Centrado horizontal cabecera
#             ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),                # Centrado vertical cabecera
#             ('BOX', (0, 0), (-1, 0), 0.5, colors.black),          # Bordes externos cabecera
#             ('GRID', (0, 0), (-1, 0), 0.5, colors.black),         # Bordes internos cabecera

#             ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),                 # Alinear columnas numéricas
#             ('VALIGN', (0, 1), (-1, -1), 'TOP'),                  # Alineación vertical en datos
#             ('BACKGROUND', (0, 1), (-1, -1), colors.white),       # Fondo blanco filas de datos
#         ]))

#         # 📍 Dibujar tabla centrada
#     table.wrapOn(pdf, x_position, 500)
#     table.drawOn(pdf, x_position, y_position)

#         # 🔢 Totales alineados con la tabla
#     subtotal = balance_subtotal(factura)
#     total = balance_total(factura)
#     iva = total - subtotal

#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(420, 280, f"SUB-TOTAL USD:")
#     pdf.drawString(420, 265, f"I.V.A. % USD:")
#     pdf.drawString(420, 250, f"TOTAL A PAGAR USD:")

    
#     # 🔢 Valores en fuente regular
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(550, 280, f"{subtotal:.2f}")
#     pdf.drawString(550, 265, f"{iva:.2f}")
#     pdf.drawString(550, 250, f"{total:.2f}")
 

#     # Nota legal
#     pdf.setFont("Helvetica-Oblique", 9)
#     pdf.drawString(50, 250, "A los efectos de lo previsto en el Art. 25 de la ley de Impuesto al Valor Agregado,")
#     pdf.drawString(50, 237, f"para efecto de conversión se ha utilizado la tasa de cambio del BCV, el 13 de julio de 2023,")
#     pdf.drawString(50, 224, "Fuente: www.bcv.org.ve")

#     pdf.showPage()
#     pdf.save()

#     return response

# MODELO FACTURA 2
# def generate_pdf(request, pk):
#     factura = Facturas.objects.get(pk=pk)
    
#     def balance_subtotal(factura):
#         subtotal = 0
#         for transaction in factura.get_factura_transaction():
#             subtotal += transaction.calcular_subtotal()
#         return subtotal
    
#     def balance_total(factura):
#         total = 0
#         for transaction in factura.get_factura_transaction():
#             total += transaction.calcular_total()
#         return total
    

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="factura_{pk}.pdf"'
#     pdf = canvas.Canvas(response, pagesize=letter)
    

#     # Encabezado
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(300, 750, "Factura N°:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(380, 750, str(factura.invoice_n))

#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(300, 735, "Control N°:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(380, 735, str(factura.invoice_c))

#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(300, 720, "Fecha:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(350, 720, factura.invoice_d.strftime("%d/%m/%Y"))
    
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(300, 705, "Cliente:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(300, 690, str(factura.partner_id))

#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(300, 675, str(factura.partner_id.tin))

#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(300, 660, "Domicilio Fiscal:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(395, 660, str(factura.partner_id.name))

#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(300, 645, "Contactos:")
#     pdf.setFont("Helvetica", 12)
#     pdf.drawString(300, 630, str(factura.partner_id.website))

    

   
#     # Tabla de productos
#     data = [["Producto", "PRECIO", "CANTIDAD", "TOTAL"]]
#     for product in factura.get_factura_transaction():
#        data.append([
#             product.product_id,
#             f"{product.price} $",
#             product.qty,
#             f"{product.calcular_subtotal():.2f} $",
           
#         ])

#     # 📐 Ancho total para centrar
#     colWidths = [200, 130, 130, 115]  # Total: 595
#     # x_position = (page_width - total_width) / 2  # 42.5 → puedes redondear a 42
#     x_position = 20
#     y_position = 480

#         # 🧾 Construcción y estilo de tabla
#     table = Table(data, colWidths=colWidths)
#     table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (-1, 0), colors.white),
#             ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
#             ('FONTSIZE', (0, 0), (-1, 0), 9),  # Tamaño de fuente más pequeño en el encabezado
#             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),      # Negrita
#             ('ALIGN', (0, 0), (-1, 0), 'CENTER'),                 # Centrado horizontal cabecera
#             ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),                # Centrado vertical cabecera
#             ('BOX', (0, 0), (-1, 0), 0.5, colors.black),          # Bordes externos cabecera
#             ('GRID', (0, 0), (-1, 0), 0.5, colors.black),         # Bordes internos cabecera

#             ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),                 # Alinear columnas numéricas
#             ('VALIGN', (0, 1), (-1, -1), 'TOP'),                  # Alineación vertical en datos
#             ('BACKGROUND', (0, 1), (-1, -1), colors.white),       # Fondo blanco filas de datos
#         ]))

#         # 📍 Dibujar tabla centrada
#     table.wrapOn(pdf, x_position, 500)
#     table.drawOn(pdf, x_position, y_position)
    
#         # Crear el PDF
#     ancho, alto = letter
    
#     # Calcular valores
#     subtotal = balance_subtotal(factura)
#     total = balance_total(factura)
#     iva = total - subtotal
#     tasa = 150.00  # Ejemplo de tasa de cambio, reemplaza con el valor real
    
#     # Calcular valores en bolívares
#     subtotal_bs = subtotal * Decimal(str(tasa))
#     iva_bs = iva *  Decimal(str(tasa))
#     total_bs = total *  Decimal(str(tasa))

#         # 🔢 Totales alineados con la tabla
#     subtotal = balance_subtotal(factura)
#     total = balance_total(factura)
#     iva = total - subtotal

#      # Configurar coordenadas
#     x_col1 = 50 * mm  # Primera columna
#     x_col2 = 130 * mm  # Segunda columna
    
#     y_start = 150 * mm  # Posición inicial
#     line_height = 7 * mm  # Altura entre líneas
    
#     # Dibujar recuadro
#     pdf.rect(x_col1 - 2*mm, y_start - 5*mm, 80*mm, 45*mm, stroke=1, fill=0)
    
#     # Título
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(x_col1, y_start, "RESUMEN DE PAGO")
    
#     # Línea separadora
#     pdf.line(x_col1, y_start - 2*mm, x_col1 + 70*mm, y_start - 2*mm)
    
#     # Valores en USD
#     y = y_start - line_height
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(x_col1, y, "SUB-TOTAL USD:")
#     pdf.drawString(x_col2, y, f"${subtotal:,.2f}".replace(",", "."))
    
#     y -= line_height
#     pdf.drawString(x_col1, y, "I.V.A. 16% USD:")
#     pdf.drawString(x_col2, y, f"${iva:,.2f}".replace(",", "."))
    
#     y -= line_height
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(x_col1, y, "TOTAL A PAGAR USD:")
#     pdf.drawString(x_col2, y, f"${total:,.2f}".replace(",", "."))
    
#     # Información de tasa de cambio
#     y -= line_height * 1.5
#     pdf.setFont("Helvetica-Oblique", 8)
#     pdf.drawString(x_col1, y, f"Tasa de cambio BCV (Bs/Dólar {tasa:,.4f})".replace(",", "."))
    
#     # Valores en Bs.
#     y -= line_height * 1.5
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(x_col1, y, "SUB-TOTAL Bs.:")
#     pdf.drawString(x_col2, y, f"{subtotal_bs:,.2f}".replace(",", "."))
    
#     y -= line_height
#     pdf.drawString(x_col1, y, "I.V.A. 16% Bs.:")
#     pdf.drawString(x_col2, y, f"{iva_bs:,.2f}".replace(",", "."))
    
#     y -= line_height
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(x_col1, y, "TOTAL A PAGAR Bs.:")
#     pdf.drawString(x_col2, y, f"{total_bs:,.2f}".replace(",", "."))
    
#     # Nota legal
#     y_nota = 50 * mm
#     pdf.setFont("Helvetica-Oblique", 9)
#     pdf.drawString(x_col1, y_nota, "A los solos efectos de lo previsto en el Art. 25 de la ley de Impuesto de Valor")
#     pdf.drawString(x_col1, y_nota - line_height/2, "Agregado se expresan los montos de la factura en Bolívares(Bs) considerando la tasa")
#     pdf.drawString(x_col1, y_nota - line_height, f"de cambio establecida por el BCV en fecha, fuente www.bcv.org.ve")
    
#     # Guardar PDF

#     pdf.showPage()
#     pdf.save()

#     return response

# from decimal import Decimal
# from reportlab.lib import colors
# from reportlab.lib.pagesizes import letter
# from reportlab.lib.units import mm
# from reportlab.platypus import Table, TableStyle
# from reportlab.pdfgen import canvas
# from django.http import HttpResponse

# def generate_pdf(request, pk):
#     factura = Facturas.objects.get(pk=pk)

#     # Funciones auxiliares para cálculo
#     def balance_subtotal(factura):
#         return sum(tx.calcular_subtotal() for tx in factura.get_factura_transaction())

#     def balance_total(factura):
#         return sum(tx.calcular_total() for tx in factura.get_factura_transaction())

#     # Preparar respuesta PDF
#     response = HttpResponse(content_type='application/pdf') 
#     response['Content-Disposition'] = f'attachment; filename="factura_{pk}.pdf"'
#     pdf = canvas.Canvas(response, pagesize=letter)
#     ancho, alto = letter

#     # 1. ENCABEZADO
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(40, 750, "ASAP SERVICIOS, C.A.")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(40, 735, "RIF: J-103997390")
#     pdf.drawString(40, 720, "Av. Francisco de Miranda, Edif. Parque Cristal Torre Oeste")
#     pdf.drawString(40, 705, "Piso 12, Oficina 12-B, Los Palos Grandes, Caracas, 1060")
#     pdf.drawString(40, 690, "Tel: 0212-2867222 / 0226-7332853861")

#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(350, 750, f"Factura No: {factura.invoice_n or '00008'}")
#     pdf.drawString(350, 735, f"Fecha: {factura.invoice_d.strftime('%d/%m/%Y') or '17/05/2022'}")

#     # 2. DATOS DEL CLIENTE
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawString(40, 660, "CLIENTE:")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(110, 660, factura.partner_id.name or "ASAP SERVICIOS, C.A.")
#     pdf.drawString(110, 645, f"RIF: {factura.partner_id.tin or 'J103997390'}")
  

#     # 3. TABLA DE ITEMS
#     items = [
#         ["Item", "Descripción", "Cantidad", "Precio Unit.", "Total"],
#     ]
#     # Ejemplo de hardcode si no vienes de BD:
#     items_data = [
#         ["01", "Mantenimiento anual (01/04/2022 al 31/03/2023) sistema STIPENDA para 300 col.", "1", "$3.480,00", "$3.480,00"],
#         ["02", "Bloque de horas de Consultoría, Asesoría o Desarrollos",           "1", "$1.000,00", "$1.000,00"],
#         ["03", "Bloque de horas de Consultoría, Asesoría o Desarrollos",           "1", "$1.000,00", "$1.000,00"],
#         ["04", "Bloque de horas de Consultoría, Asesoría o Desarrollos",           "1", "$1.000,00", "$1.000,00"],
#     ]
#     for row in items_data:
#         items.append(row)

#     colWidths = [30*mm, 100*mm, 25*mm, 30*mm, 30*mm]
#     table = Table(items, colWidths=colWidths)
#     table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
#         ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('ALIGN',      (2, 1), (-1, -1), 'CENTER'),
#         ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
#         ('FONTSIZE',   (0, 0), (-1, -1), 9),
#         ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
#     ]))
#     table.wrapOn(pdf, 40, 500)
#     table.drawOn(pdf, 40, 480)

#     # 4. RESUMEN DE PAGOS
   
# # 4. CÁLCULOS ESPECÍFICOS PARA EL RESUMEN
#     subtotal_usd = Decimal("3480.00")
#     iva_usd      = (subtotal_usd * Decimal("0.16")).quantize(Decimal("0.01"))
#     total_usd    = (subtotal_usd + iva_usd).quantize(Decimal("0.01"))
#     tasa         = Decimal("4.7848")

#     subtotal_bs  = (subtotal_usd * tasa).quantize(Decimal("0.01"))
#     iva_bs       = (iva_usd      * tasa).quantize(Decimal("0.01"))
#     total_bs     = (total_usd    * tasa).quantize(Decimal("0.01"))

#     # 5. DIBUJAR EL RECUADRO DE RESUMEN
#     x = 40 * mm
#     y_top = 160 * mm       # coordenada Y del tope del recuadro
#     box_w = 80 * mm
#     box_h = 60 * mm

#     # rectángulo externo
#     pdf.rect(x, y_top - box_h, box_w, box_h, stroke=1, fill=0)


#     # líneas de valores
#     pdf.setFont("Helvetica", 9)
#     line_gap = 6 * mm
#     current_y = y_top - 15 * mm

#     # USD
#     pdf.drawString(x + 3*mm, current_y, "SUB-TOTAL:")
#     pdf.drawRightString(x + box_w - 3*mm, current_y, f"${subtotal_usd:,.2f}".replace(",", "."))

#     current_y -= line_gap
#     pdf.drawString(x + 3*mm, current_y, "I.V.A. 16%:")
#     pdf.drawRightString(x + box_w - 3*mm, current_y, f"${iva_usd:,.2f}".replace(",", "."))

#     current_y -= line_gap
#     pdf.setFont("Helvetica-Bold", 9)
#     pdf.drawString(x + 3*mm, current_y, "TOTAL A PAGAR:")
#     pdf.drawRightString(x + box_w - 3*mm, current_y, f"${total_usd:,.2f}".replace(",", "."))

#     # tipo de cambio
#     current_y -= (line_gap + 2*mm)
#     pdf.setFont("Helvetica-Oblique", 7)
#     pdf.drawString(x + 3*mm, current_y, f"Tasa BCV (17/05/2022): Bs. {tasa:.4f}")

#     # Bs
#     pdf.setFont("Helvetica", 9)
#     current_y -= (line_gap + 1*mm)
#     pdf.drawString(x + 3*mm, current_y, "SUB-TOTAL Bs.:")
#     pdf.drawRightString(x + box_w - 3*mm, current_y, f"{subtotal_bs:,.2f}".replace(",", "."))

#     current_y -= line_gap
#     pdf.drawString(x + 3*mm, current_y, "I.V.A. 16% Bs.:")
#     pdf.drawRightString(x + box_w - 3*mm, current_y, f"{iva_bs:,.2f}".replace(",", "."))

#     current_y -= line_gap
#     pdf.setFont("Helvetica-Bold", 9)
#     pdf.drawString(x + 3*mm, current_y, "TOTAL A PAGAR Bs.:")
#     pdf.drawRightString(x + box_w - 3*mm, current_y, f"{total_bs:,.2f}".replace(",", "."))

#     # nota legal
#     nota_y = y_top - box_h + 5*mm
#     pdf.setFont("Helvetica-Oblique", 7)
#     pdf.drawString(x + 3*mm, nota_y,   "A los solos efectos de lo previsto en el Art. 25 de la ley de IVA,")
#     pdf.drawString(x + 3*mm, nota_y - 3*mm,
#                 "se expresan los montos en Bolívares según la tasa del BCV del 17/05/2022.")



#         # Finalizar y devolver
#     pdf.showPage()
#     pdf.save()
#     return response

#MODELO FACTURA 2 DEEP SEEK
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.platypus import Table, TableStyle
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from reportlab.lib import colors

def generate_pdf(request, pk):
    factura = Facturas.objects.get(pk=pk)
    verde_73A243 = colors.Color(red=170/255, green=200/255, blue=130/255)

    # Funciones auxiliares para cálculo
    def balance_subtotal(factura):
        return sum(tx.calcular_subtotal() for tx in factura.get_factura_transaction())

    def balance_total(factura):
        return sum(tx.calcular_total() for tx in factura.get_factura_transaction())

    # Preparar respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="factura_{pk}.pdf"'
    pdf = canvas.Canvas(response, pagesize=letter)
    ancho, alto = letter

    # 1. ENCABEZADO
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(40, 750, "ASAP SERVICIOS, C.A.")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, 735, "RIF: J-103997390")
    pdf.drawString(40, 720, "Av. Francisco de Miranda, Edif. Parque Cristal Torre Oeste")
    pdf.drawString(40, 705, "Piso 12, Oficina 12-B, Los Palos Grandes, Caracas, 1060")
    pdf.drawString(40, 690, "Tel: 0212-2867222 / 0226-7332853861")

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(350, 750, f"Factura No: {factura.invoice_n or '00008'}")
    pdf.drawString(350, 735, f"Fecha: {factura.invoice_d.strftime('%d/%m/%Y') or '17/05/2022'}")

    # 2. DATOS DEL CLIENTE
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(40, 660, "CLIENTE:")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(110, 660, factura.partner_id.name or "ASAP SERVICIOS, C.A.")
    pdf.drawString(110, 645, f"RIF: {factura.partner_id.tin or 'J103997390'}")

    # 3. TABLA DE ITEMS
    items = [
        ["Item", "Descripción", "Cantidad", "Precio Unit.", "Total"],
    ]
    
    # Obtener datos de la factura o usar datos de ejemplo
    if factura.get_factura_transaction().exists():
        for i, transaction in enumerate(factura.get_factura_transaction(), 1):
            items.append([
                f"{i:02d}",
                str(transaction.product_id),
                str(transaction.qty),
                f"${transaction.price:,.2f}".replace(",", "."),
                f"${transaction.calcular_subtotal():,.2f}".replace(",", ".")
            ])
    else:
        # Datos de ejemplo si no hay transacciones
        items_data = [
            ["01", "Mantenimiento anual (01/04/2022 al 31/03/2023) sistema STIPENDA para 300 col.", "1", "$3.480,00", "$3.480,00"],
            ["02", "Bloque de horas de Consultoría, Asesoría o Desarrollos", "4", "$0,00", "$0,00"],
            ["03", "Bloque de horas de Consultoría, Asesoría o Desarrollos", "2", "$0,00", "$0,00"],
            ["04", "Bloque de horas de Consultoría, Asesoría o Desarrollos", "40", "$0,00", "$0,00"],
        ]
        for row in items_data:
            items.append(row)

    colWidths = [30*mm, 100*mm, 25*mm, 30*mm, 30*mm]
    table = Table(items, colWidths=colWidths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    table.wrapOn(pdf, 40, 500)
    table.drawOn(pdf, 40, 480)

    # 4. CÁLCULOS ESPECÍFICOS PARA EL RESUMEN
    # Usar datos reales o de ejemplo
    if factura.get_factura_transaction().exists():
        subtotal_usd = balance_subtotal(factura)
        iva_usd = (subtotal_usd * Decimal("0.16")).quantize(Decimal("0.01"))
        total_usd = (subtotal_usd + iva_usd).quantize(Decimal("0.01"))
    else:
        subtotal_usd = Decimal("3480.00")
        iva_usd = (subtotal_usd * Decimal("0.16")).quantize(Decimal("0.01"))
        total_usd = (subtotal_usd + iva_usd).quantize(Decimal("0.01"))
    
    tasa = Decimal("4.7848")
    subtotal_bs = (subtotal_usd * tasa).quantize(Decimal("0.01"))
    iva_bs = (iva_usd * tasa).quantize(Decimal("0.01"))
    total_bs = (total_usd * tasa).quantize(Decimal("0.01"))

    # 5. DIBUJAR EL RECUADRO DE RESUMEN (ESTILO TABLA)
    # Posicionar el recuadro en la parte inferior derecha
    x = 420  # Posición X (desde la izquierda)
    y = 171  # Posición Y (desde abajo)
    
    # Crear tabla para el resumen
    summary_data = [
        ["SUB-TOTAL", f"${subtotal_usd:,.2f}".replace(",", ".")],
        ["I.V.A. 16%", f"${iva_usd:,.2f}".replace(",", ".")],
        ["TOTAL A PAGAR", f"${total_usd:,.2f}".replace(",", ".")],
    ]
    
    summary_table = Table(summary_data, colWidths=[80, 60])
    summary_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica'),
        ('BACKGROUND', (0, 0), (0, 2), verde_73A243),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 2), (1, 2), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LINEABOVE', (0, 2), (1, 2), 1, colors.black),
    ]))
    
    summary_table.wrapOn(pdf, x, y)
    summary_table.drawOn(pdf, x, y)
    
       # Información de tasa de cambio ENCERRADA EN UN RECTÁNGULO
    tasa_text = f"Tasa de cambio establecida por el Banco Central de Venezuela BCV (Bs/Dólar {tasa:.4f})"
    pdf.setFont("Helvetica", 9)
    
    # Calcular el ancho del texto para dimensionar el rectángulo
    text_width = pdf.stringWidth(tasa_text, "Helvetica", 9)
    
    # Dibujar rectángulo alrededor del texto
    rect_x =35  # Ajustar posición X con margen
    rect_y = 162 - 12  # Ajustar posición Y con margen
    rect_width = text_width + 176.9  # Ancho del texto + margen
    rect_height = 21  # Alto del rectángulo
    
    pdf.rect(rect_x, rect_y, rect_width, rect_height)
    
    # Dibujar el texto centrado dentro del rectángulo
    pdf.drawString(150, 160, tasa_text)
    
    # Nota legal
    pdf.setFont("Helvetica", 9)
    nota_text = "A los solos efectos de lo previsto en el Art. 25 de la ley de Impuesto de Valor"
    pdf.drawString(40, y - 35, nota_text)
    nota_text2 = "Agregado se expresan los montos de la factura en Bolívares(Bs) considerando la tasa"
    pdf.drawString(40, y - 45, nota_text2)
    nota_text3 = "de cambio establecida por el BCV en fecha 17/05/2022, fuente www.bcv.org.ve"
    pdf.drawString(40, y - 55, nota_text3)
    
    # Resumen en bolívares
    summary_bs_data = [
        ["SUB-TOTAL", f"Bs. {subtotal_bs:,.2f}".replace(",", ".")],
        ["I.V.A. 16%", f"Bs. {iva_bs:,.2f}".replace(",", ".")],
        ["TOTAL A PAGAR", f"Bs. {total_bs:,.2f}".replace(",", ".")],
    ]
    
    summary_bs_table = Table(summary_bs_data, colWidths=[80, 60])
    summary_bs_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica'),
        ('BACKGROUND', (0, 0), (0, 2), verde_73A243),   
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 2), (1, 2), 'Helvetica-Bold'), 
        ('LINEABOVE', (0, 2), (1, 2), 1, colors.black),
    ]))
    
    summary_bs_table.wrapOn(pdf, x, y - 75)
    summary_bs_table.drawOn(pdf, x, y - 75)

    # Finalizar y devolver
    pdf.showPage()
    pdf.save()
    return response

# MODELO FACTURA 3
# from reportlab.lib.pagesizes import LETTER
# from reportlab.pdfgen import canvas
# from reportlab.lib import colors
# from reportlab.platypus import Table, TableStyle
# from reportlab.lib.styles import ParagraphStyle
# from reportlab.platypus import Paragraph
# from reportlab.lib.units import inch
# from datetime import timedelta
# from decimal import Decimal


# def generate_pdf(request, pk):
#     factura = Facturas.objects.get(pk=pk)
#     transacciones = factura.get_factura_transaction() 

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="factura_{pk}.pdf"'
#     pdf = canvas.Canvas(response, pagesize=LETTER)
#     width, height = LETTER

#     # Set up fonts and colors
#     pdf.setTitle(f"Factura NetUno {factura.invoice_n}")
    
#     # Header with NetUno information
#     pdf.setFont("Helvetica-Bold", 8)
#     pdf.drawString(50, height - 30, "Número C.A. 3-30308196")
#     pdf.drawString(50, height - 42, "Calle 7 Gallimardia II Piso 2 OI N/A Uro La Unicia")
#     pdf.drawString(50, height - 54, "Ciencias (Paseo) Maraca zona Postal 1075")
#     pdf.drawString(50, height - 66, "Teléfono: 002157 7120-06")
#     pdf.drawString(50, height - 78, "Contact Center: 059 0800-900-900")
#     pdf.drawString(50, height - 90, "Código de Autoridad: 6190")

#     # Client information
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(50, height - 120, "Nombre / Razón Social:")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(180, height - 120, str(factura.partner_id.name))
    
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(50, height - 135, "C.I. / R.I.F.:")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(180, height - 135, str(factura.partner_id.tin))
    
   
  

#     # Invoice header
#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.drawCentredString(width / 2, height - 180, "Factura")
    
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(50, height - 200, "Nro. de Factura:")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(150, height - 200, str(factura.invoice_n))
    
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(250, height - 200, "Fecha de Emisión:")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(350, height - 200, factura.invoice_d.strftime("%d/%m/%Y"))
    
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(50, height - 215, "Nro. de Control:")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(150, height - 215, str(factura.invoice_c))
    
#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(250, height - 215, "Fecha de Asignación:")
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(380, height - 215, (factura.invoice_d + timedelta(days=1)).strftime("%d/%m/%Y"))

#     # Account and payment information
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(50, height - 240, f"Número de Cuenta: 1276116")
#     pdf.drawString(400, height - 240, f"Pague Antes de: {(factura.invoice_d + timedelta(days=17)).strftime('%d-%m-%Y')}")

#     # Table of items
#     data = [
#         ["Cantidad", "Código", "Descripción", "Precio Unitario", "Total"]
#     ]
    
#     for trans in transacciones:
#         precio = trans.price  # Precio unitario original
#         iva_percent = Decimal('0.16')
#         iva_amount = (precio * Decimal(trans.qty) * iva_percent).quantize(Decimal('0.00'))
#         total_linea = (precio * Decimal(trans.qty)).quantize(Decimal('0.00'))
        
#         data.append([
#             str(trans.qty),
#             trans.product_id or "N/A",
#             trans.product_id.name,
#             f"$ {precio:.2f}",  # Precio unitario original
#             f"$ {total_linea:.2f}"  # Total sin IVA incluido
#         ])
#     colWidths = [50, 50, 200, 60, 50, 50, 60]
#     table = Table(data, colWidths=colWidths)
    
#     table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4472C4")),
#         ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#         ('FONTSIZE', (0, 0), (-1, -1), 8),
#         ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
#         ('BACKGROUND', (0, 1), (-1, -1), colors.white),
#         ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
#         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#         ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Alinea precios, IVA y total a la derecha
#     ]))
    
#     table.wrapOn(pdf, 50, height - 400)
#     table.drawOn(pdf, 50, height - 400)

#     # Totals section
#     subtotal = sum(Decimal(str(trans.price * trans.qty)) for trans in transacciones) 
#     iva_total = subtotal * Decimal("0.16")
#     total = subtotal + iva_total
#     exchange_rate = Decimal(97.3126)  # BCV rate

#     pdf.setFont("Helvetica-Bold", 10)
#     pdf.drawString(350, height - 430, "Total Ventas no Gravadas:")
#     pdf.drawString(350, height - 445, "Total Base imponible del 16,00%:")
#     pdf.drawString(350, height - 460, "Total IVA 16,00%:")
#     pdf.drawString(350, height - 475, "Total Factura:")
    
#     pdf.setFont("Helvetica", 10)
#     pdf.drawString(530, height - 430, "Bs. 0,00")
#     pdf.drawString(530, height - 445, f"Bs. {subtotal * exchange_rate:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
#     pdf.drawString(530, height - 460, f"Bs. {iva_total * exchange_rate:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
#     pdf.drawString(530, height - 475, f"Bs. {total * exchange_rate:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

#     # Legal notice about IGTF
#     pdf.setFont("Helvetica", 8)
#     pdf.drawString(50, height - 500, "Los montos pagados en divisas estarán sujetos al 3% del IGTF de acuerdo a la Gaceta Oficial N 42.339 de fecha 17-03-2022")

#     # NetUno message
#     message = """<b>Mensaje de Facturación:</b> Nacímos hace 30 años, con una promesa. Hemos estado aquí, impulsando sueños, mejorando la vida en cada hogar y en cada empresa. Porque los momentos que importan no se cuentan en gigas, canales o llamadas, sino en lo que hace que tu corazón vibre con más fuerza. Somos más que datos, somos progreso que conecta. NetUno, 30 años contigo en el corazón."""
    
#     style = ParagraphStyle(
#         name='Normal',
#         fontName='Helvetica',
#         fontSize=9,
#         leading=12,
#         spaceBefore=6,
#         spaceAfter=6,
#     )
    
#     p = Paragraph(message, style)
#     p.wrapOn(pdf, width - 100, 100)
#     p.drawOn(pdf, 50, height - 550)

#     # Footer with NetUno branding
#     pdf.setFont("Helvetica-Bold", 12)
#     pdf.drawCentredString(width / 2, height - 580, "Nacímos con la promesa")
#     pdf.drawCentredString(width / 2, height - 595, "DECONECTARNOS CONTIGO")
    
#     pdf.setFont("Helvetica-Bold", 14)
#     pdf.drawCentredString(width / 2, height - 615, "NETUNO")
#     pdf.drawCentredString(width / 2, height - 630, "30 AÑOS CONTIGO EN EL CORAZÓN")

#     # Legal footer
#     pdf.setFont("Helvetica", 6)
#     pdf.drawString(50, height - 650, "Corporación Unidigital 1220, C.A. Rif J-40148330-5 Imprenta Digital, Autorizada según Providencia Administrativa SEMIAT/INTI/2021 0000001 de fecha 19-01-2021,")
#     pdf.drawString(50, height - 660, f"Numero de Control desde el Nro 00-10749176 hasta el Nro 00-11049175 generadas el {factura.invoice_d.strftime('%d-%m-%Y')}")

#     # Exchange rate
#     pdf.setFont("Helvetica-Bold", 8)
#     pdf.drawString(50, height - 680, "Tasa de Cambio BCV:")
#     pdf.drawString(150, height - 680, f"Bs. {exchange_rate:,.4f}".replace(",", "X").replace(".", ",").replace("X", "."))

#     pdf.showPage()
#     pdf.save()

#     return response


    #generar Excel
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side



class ReportePersonalizadoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        pk = self.kwargs.get('pk')
        factura = Facturas.objects.get(id=pk)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte'

        # Convertir el objeto User a cadena de texto
        # user_str = str(factura.usuario)  # Asumiendo que 'usuario' es un campo en Factura
        # ws.append([user_str, factura.descripcion, factura.total])
        
        
            # empezamos a configurar la hoja de excel con el contenido
            
             #titulo
        ws["B1"].alignment = Alignment(horizontal="left", vertical="center") #modificamos la alineacion
        # ws["B1"].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                              top = Side(border_style="thin"), bottom = Side(border_style="thin")) #modificamos el borde
            
            #Relleno de la celda
        ws["B1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            #Aplicamos fuente
        ws["B1"].font = Font(name='Times New Roman', size=12, bold=True)
            
            #asignamos valor a la celda B1
        ws["B1"] = "REPORTE  DE FACTURAS PERZONALIZADO EN EXCEL CON DJANGO"
            
            #combinamos las celdas
        ws.merge_cells("B1:E1")
            
            #cmbiar anchura de la columna
        ws.column_dimensions["B"].width=35
        ws.column_dimensions["C"].width=35
        ws.column_dimensions["D"].width=35
        ws.column_dimensions["E"].width=35
            
            #cambiar el tamano de las filas
        ws.row_dimensions[1].height = 25
            
        #   crear la cabecera
       # Ajustar la alineación a la izquierda
        ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

        # Mantener el borde
        # ws["B3"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B3"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B3'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B3'] = 'CLIENTE:'
           
            
        ws["B7"].alignment = Alignment(horizontal="left", vertical="center")

        # # Mantener el borde
        # ws["B7"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B7"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B7'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B7'] = 'NÚMERO DE FACTURA:'
        
       
            
         # Ajustar la alineación a la izquierda
        ws["B11"].alignment = Alignment(horizontal="left", vertical="center")

        # # Mantener el borde
        # ws["B11"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B11"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B11'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B11'] = 'R.I.F:'
        
        
         # Ajustar la alineación a la izquierda
        ws["B15"].alignment = Alignment(horizontal="left", vertical="center")

        # # Mantener el borde
        # ws["B15"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B15"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B15'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B15'] = 'DOMICILIO FISCAL:'
        
       
         # Ajustar la alineación a la izquierda
        ws["B19"].alignment = Alignment(horizontal="left", vertical="center")

        # # Mantener el borde
        # ws["B19"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B19"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B19'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B19'] = 'TELEFONO:'
        
         # Ajustar la alineación a la izquierda
        ws["B23"].alignment = Alignment(horizontal="left", vertical="center")

        # # Mantener el borde
        # ws["B23"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B23"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B23'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B23'] = 'DESCRIPCIÓN:'
       
         # Ajustar la alineación a la izquierda
        ws["B27"].alignment = Alignment(horizontal="left", vertical="center")

        # # Mantener el borde
        # ws["B27"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B27"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B27'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B27'] = 'FORMA DE PAGO:'
        
         # Ajustar la alineación a la izquierda
        ws["B31"].alignment = Alignment(horizontal="left", vertical="center")

        # # Mantener el borde
        # ws["B31"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B31"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B31'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B31'] = 'BASE IMPONIBLE:'
        
         # Ajustar la alineación a la izquierda
        ws["B35"].alignment = Alignment(horizontal="left", vertical="center")

        # Mantener el borde
        # ws["B35"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B35"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B35'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B35'] = 'I.V.A (%) Bs:'
    
        
          # Ajustar la alineación a la izquierda
        ws["B39"].alignment = Alignment(horizontal="left", vertical="center")

        # Mantener el borde
        # ws["B39"].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
        #                         top=Side(border_style="thin"), bottom=Side(border_style="thin"))

        # Cambiar el fondo a blanco
        ws["B39"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cambiar la fuente a Times New Roman
        ws['B39'].font = Font(name='Times New Roman', size=12, bold=True)

        # Asignar el valor a la celda
        ws['B39'] = 'TOTAL:'
        
        
      
        
        # # Combinar las celdas de las columnas C y D
         # Definir el rango de filas que deseas combinar
        start_row = 1
        end_row = 40

        for row in range(start_row, end_row + 1):
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            
          #pintamos los datos
            
            
        ws.cell(row = 3, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row = 3, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 3, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 3 , column = 3).value = str(factura.cliente)
            
        ws.cell(row = 7, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row = 7, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 7, column = 3).font = Font(name='Times New Roman', size=12)
        factura.save()
        ws.cell(row = 7, column = 3).value = factura.numero_factura
        ws.merge_cells("C3:D3")
            
        ws.cell(row = 11, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row =11, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 11, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 11, column = 3).value = factura.rif
        
        
        ws.cell(row = 15, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row =15, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 15, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 15, column = 3).value = factura.domicilio_fiscal
        
        
        
        ws.cell(row = 19, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row =19, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 19, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 19, column = 3).value = factura.telefono
        
        
        
        ws.cell(row = 23, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row =23, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 23, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 23, column = 3).value = factura.descripcion
        
        
        ws.cell(row = 27, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row =27, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 27, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 27, column = 3).value = factura.forma_pago
        
        
        
        ws.cell(row = 31, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row =31, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 31, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 31, column = 3).value = str(factura.importe) + "Bs"
        
        
        
        ws.cell(row = 35, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row =35, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 35, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 35, column = 3).value = str(factura.iva) + "%"
        
        
        
        ws.cell(row = 39, column = 3).alignment = Alignment(horizontal="left", vertical="center") #cell son celdas
        # ws.cell(row =39, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
        #                             top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row = 39, column = 3).font = Font(name='Times New Roman', size=12)
        ws.cell(row = 39, column = 3).value = str(factura.calcular_total_con_iva()) + "Bs"
        
        


        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_facturas.xlsx"'
        wb.save(response)
        return response

        
        
        
        
        
        
        
        
        
        # campo = int(request.GET.get('campo')) #recibe la respuesta del form y lo convierte a int
#         pk = self.kwargs.get('pk')
#         factura = Factura.objects.filter(id = pk) #solo traes los que coincidan con la edad
#         # query = Factura.objects.all()
#          # Convertir el objeto User a cadena de texto
#         user_str = str(factura.usuario)  # Asumiendo que 'usuario' es un campo en Factura
#         ws.append([user_str, factura.descripcion, factura.total])
#         wb = Workbook()
#         bandera = True #hay que hacer una validacion para que no deje la hoja en blanco cuando creemos mas hojas
#         cont = 1
#         controlador = 4 #es la posicion inicial para pintar datos
#         # for q in query: #esto es para cargar cada query en cada hoja 
#         if bandera:
#                 ws = wb.active #te permite acceder a la hoja de trabajo activa
#                 ws.title = 'Hoja'+str(cont) #le damos el nombre a la primera hoja activa
#                 bandera = False
#             # else:
#             #     ws = wb.create_sheet('Hoja'+str(cont)) #creamos otra hoja de trabajo siguiente
           
            
            
#             #empezamos a configurar la hoja de excel con el contenido
            
#              #titulo
#         ws["B1"].alignment = Alignment(horizontal="center", vertical = "center") #modificamos la alineacion
#         ws["B1"].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
#                                      top = Side(border_style="thin"), bottom = Side(border_style="thin")) #modificamos el borde
            
#             #Relleno de la celda
#         ws["B1"].fill = PatternFill(start_color="66FFCC", end_color = "66FFCC", fill_type = "solid")
            
#             #Aplicamos fuente
#         ws["B1"].font = Font(name = "Calibri", size = 12, bold = True)
            
#             #asignamos valor a la celda B1
#         ws["B1"] = "REPORTE  DE FACTURAS PERZONALIZADO EN EXCEL CON DJANGO"
            
#             #combinamos las celdas
#         ws.merge_cells("B1:E1")
            
#             #cmbiar anchura de la columna
#         ws.column_dimensions["B"].width=20
#         ws.column_dimensions["C"].width=20
#         ws.column_dimensions["D"].width=20
#         ws.column_dimensions["E"].width=20
            
#             #cambiar el tamano de las filas
            
#         ws.row_dimensions[1].height = 25
            
#             #crear la cabecera
#         ws["B3"].alignment = Alignment(horizontal="center", vertical = "center")
#         ws["B3"].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
#                                      top = Side(border_style="thin"), bottom = Side(border_style="thin"))
#         ws["B3"].fill = PatternFill(start_color="66CFCC", end_color = "66CFCC", fill_type = "solid")
#         ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
#         ws['B3'] = 'Usuario'
           
            
#         ws["C3"].alignment = Alignment(horizontal="center", vertical = "center")
#         ws["C3"].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
#                                      top = Side(border_style="thin"), bottom = Side(border_style="thin"))
#         ws["C3"].fill = PatternFill(start_color="66CFCC", end_color = "66CFCC", fill_type = "solid")
#         ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
#         ws['C3'] = 'Descripcion'
#         ws.merge_cells("C3:D3")
            
            
#         ws["E3"].alignment = Alignment(horizontal="center", vertical = "center")
#         ws["E3"].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
#                                      top = Side(border_style="thin"), bottom = Side(border_style="thin"))
#         ws["E3"].fill = PatternFill(start_color="66CFCC", end_color = "66CFCC", fill_type = "solid")
#         ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
#         ws['E3'] = 'Total'
            
#             #pintamos los datos
            
            
#         ws.cell(row = controlador, column = 2).alignment = Alignment(horizontal = "center") #cell son celdas
#         ws.cell(row = controlador, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
#                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
#         ws.cell(row = controlador, column = 2).font = Font(name = 'Calibri', size = 8)
#         ws.cell(row = controlador, column = 2).value = user_str
            
#         ws.cell(row = controlador, column = 3).alignment = Alignment(horizontal = "center") #cell son celdas
#         ws.cell(row = controlador, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
#                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
#         ws.cell(row = controlador, column = 3).font = Font(name = 'Calibri', size = 8)
#         ws.cell(row = controlador, column = 3).value = factura.descripcion
            
#         ws.cell(row = controlador, column = 5).alignment = Alignment(horizontal = "center") #cell son celdas
#         ws.cell(row = controlador, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
#                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
#         ws.cell(row = controlador, column = 5).font = Font(name = 'Calibri', size = 8)
#         ws.cell(row = controlador, column = 5).value = int(factura.total)
            
#         controlador += 1
#         cont += 1 #le da el numero a cada pagina
            
#         # Definir el rango de filas que deseas combinar
#         start_row = 1
#         end_row = 20

# # Combinar las celdas de las columnas C y D
#         for row in range(start_row, end_row + 1):
#             ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        
  

#         #Establecer el nombre de mi archivo
#         nombre_archivo = "ReportePersonalizadoExcel.xlsx"
#         #Definir el tipo de respuesta que se va a dar
#         response = HttpResponse(content_type = "application/ms-excel")
#         contenido = "attachment; filename = {0}".format(nombre_archivo)
#         response["Content-Disposition"] = contenido
#         wb.save(response)
#         return response
    
